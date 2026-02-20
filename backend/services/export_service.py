"""
Export Service - Advanced data export functionality
Supports CSV, Excel, PDF, JSON, XML for all data types
"""
import json
import csv
import io
import zipfile
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any
from bson import ObjectId
import xml.etree.ElementTree as ET
from xml.dom import minidom

# Try to import optional dependencies
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


class ExportService:
    def __init__(self, db):
        self.db = db
    
    def _serialize_doc(self, doc: dict) -> dict:
        """Convert MongoDB document to JSON-serializable format"""
        if doc is None:
            return None
        result = {}
        for key, value in doc.items():
            if key == '_id':
                result['id'] = str(value)
            elif isinstance(value, ObjectId):
                result[key] = str(value)
            elif isinstance(value, datetime):
                result[key] = value.isoformat()
            elif isinstance(value, list):
                result[key] = [self._serialize_doc(item) if isinstance(item, dict) else item for item in value]
            elif isinstance(value, dict):
                result[key] = self._serialize_doc(value)
            else:
                result[key] = value
        return result

    async def get_devices(self, organization_id: Optional[str] = None) -> List[Dict]:
        """Get devices with optional organization filter"""
        query = {}
        if organization_id:
            query['organization_id'] = organization_id
        
        devices = []
        cursor = self.db.devices.find(query, {'_id': 0})
        async for doc in cursor:
            devices.append(self._serialize_doc(doc))
        return devices

    async def get_alerts(self, organization_id: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None) -> List[Dict]:
        """Get alerts with filters"""
        query = {}
        if date_from:
            query['created_at'] = {'$gte': datetime.fromisoformat(date_from)}
        if date_to:
            if 'created_at' not in query:
                query['created_at'] = {}
            query['created_at']['$lte'] = datetime.fromisoformat(date_to)
        
        alerts = []
        cursor = self.db.alerts.find(query, {'_id': 0}).sort('created_at', -1).limit(10000)
        async for doc in cursor:
            alerts.append(self._serialize_doc(doc))
        return alerts

    async def get_incidents(self, organization_id: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None) -> List[Dict]:
        """Get incidents with filters"""
        query = {}
        if date_from:
            query['created_at'] = {'$gte': datetime.fromisoformat(date_from)}
        if date_to:
            if 'created_at' not in query:
                query['created_at'] = {}
            query['created_at']['$lte'] = datetime.fromisoformat(date_to)
        
        incidents = []
        cursor = self.db.incidents.find(query, {'_id': 0}).sort('created_at', -1).limit(10000)
        async for doc in cursor:
            incidents.append(self._serialize_doc(doc))
        return incidents

    async def get_logs(self, date_from: Optional[str] = None, date_to: Optional[str] = None) -> List[Dict]:
        """Get access logs with filters"""
        query = {}
        if date_from:
            query['timestamp'] = {'$gte': datetime.fromisoformat(date_from)}
        if date_to:
            if 'timestamp' not in query:
                query['timestamp'] = {}
            query['timestamp']['$lte'] = datetime.fromisoformat(date_to)
        
        logs = []
        cursor = self.db.access_logs.find(query, {'_id': 0}).sort('timestamp', -1).limit(10000)
        async for doc in cursor:
            logs.append(self._serialize_doc(doc))
        return logs

    async def get_users(self) -> List[Dict]:
        """Get users (without passwords)"""
        users = []
        cursor = self.db.users.find({}, {'_id': 0, 'password': 0, 'hashed_password': 0})
        async for doc in cursor:
            users.append(self._serialize_doc(doc))
        return users

    async def get_organizations(self) -> List[Dict]:
        """Get organizations and groups"""
        orgs = []
        cursor = self.db.organizations.find({}, {'_id': 0})
        async for doc in cursor:
            orgs.append(self._serialize_doc(doc))
        
        groups = []
        cursor = self.db.groups.find({}, {'_id': 0})
        async for doc in cursor:
            groups.append(self._serialize_doc(doc))
        
        return {'organizations': orgs, 'groups': groups}

    # ==================== CSV Export ====================
    def to_csv(self, data: List[Dict], filename_prefix: str = "export") -> io.BytesIO:
        """Convert data to CSV"""
        output = io.StringIO()
        if not data:
            output.write("No data\n")
            return io.BytesIO(output.getvalue().encode('utf-8-sig'))
        
        # Flatten nested dicts for CSV
        flat_data = []
        for item in data:
            flat_item = {}
            for key, value in item.items():
                if isinstance(value, (dict, list)):
                    flat_item[key] = json.dumps(value, ensure_ascii=False)
                else:
                    flat_item[key] = value
            flat_data.append(flat_item)
        
        writer = csv.DictWriter(output, fieldnames=flat_data[0].keys())
        writer.writeheader()
        writer.writerows(flat_data)
        
        return io.BytesIO(output.getvalue().encode('utf-8-sig'))

    # ==================== JSON Export ====================
    def to_json(self, data: Any) -> io.BytesIO:
        """Convert data to JSON"""
        output = json.dumps(data, ensure_ascii=False, indent=2, default=str)
        return io.BytesIO(output.encode('utf-8'))

    # ==================== XML Export ====================
    def to_xml(self, data: Any, root_name: str = "data") -> io.BytesIO:
        """Convert data to XML"""
        def dict_to_xml(parent, d):
            if isinstance(d, dict):
                for key, value in d.items():
                    # Sanitize key for XML
                    safe_key = ''.join(c if c.isalnum() or c == '_' else '_' for c in str(key))
                    if safe_key[0].isdigit():
                        safe_key = '_' + safe_key
                    child = ET.SubElement(parent, safe_key)
                    dict_to_xml(child, value)
            elif isinstance(d, list):
                for item in d:
                    child = ET.SubElement(parent, 'item')
                    dict_to_xml(child, item)
            else:
                parent.text = str(d) if d is not None else ''
        
        root = ET.Element(root_name)
        if isinstance(data, list):
            for item in data:
                child = ET.SubElement(root, 'record')
                dict_to_xml(child, item)
        else:
            dict_to_xml(root, data)
        
        xml_str = ET.tostring(root, encoding='unicode')
        pretty_xml = minidom.parseString(xml_str).toprettyxml(indent="  ")
        
        return io.BytesIO(pretty_xml.encode('utf-8'))

    # ==================== Excel Export ====================
    def to_excel(self, data: List[Dict], sheet_name: str = "Data") -> io.BytesIO:
        """Convert data to Excel"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if not data:
            ws.append(["No data"])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output
        
        # Header style
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Write headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, item in enumerate(data, 2):
            for col_idx, key in enumerate(headers, 1):
                value = item.get(key, '')
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ==================== PDF Export ====================
    def to_pdf(self, data: List[Dict], title: str = "Export") -> io.BytesIO:
        """Convert data to PDF"""
        if not PDF_AVAILABLE:
            raise ImportError("reportlab is required for PDF export")
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, spaceAfter=20)
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        if not data:
            elements.append(Paragraph("No hay datos para mostrar", styles['Normal']))
        else:
            # Prepare table data
            headers = list(data[0].keys())[:8]  # Limit columns for PDF
            table_data = [headers]
            
            for item in data[:100]:  # Limit rows for PDF
                row = []
                for key in headers:
                    value = item.get(key, '')
                    if isinstance(value, (dict, list)):
                        value = '...'
                    value = str(value)[:30]  # Truncate long values
                    row.append(value)
                table_data.append(row)
            
            # Create table
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ]))
            elements.append(table)
            
            if len(data) > 100:
                elements.append(Spacer(1, 10))
                elements.append(Paragraph(f"Mostrando 100 de {len(data)} registros", styles['Normal']))
        
        doc.build(elements)
        output.seek(0)
        return output

    # ==================== Complete Export ====================
    async def export_complete(self, format: str, organization_id: Optional[str] = None) -> io.BytesIO:
        """Export all data in a single file/archive"""
        data = {
            'devices': await self.get_devices(organization_id),
            'alerts': await self.get_alerts(organization_id),
            'incidents': await self.get_incidents(organization_id),
            'users': await self.get_users(),
            'organizations': await self.get_organizations(),
            'export_date': datetime.now().isoformat(),
            'export_type': 'complete'
        }
        
        if format == 'json':
            return self.to_json(data)
        elif format == 'xml':
            return self.to_xml(data, 'siempria_backup')
        elif format == 'excel':
            # Create multi-sheet Excel
            if not EXCEL_AVAILABLE:
                raise ImportError("openpyxl required")
            
            wb = openpyxl.Workbook()
            
            # Devices sheet
            self._add_sheet(wb, 'Dispositivos', data['devices'])
            # Alerts sheet
            self._add_sheet(wb, 'Alertas', data['alerts'])
            # Incidents sheet
            self._add_sheet(wb, 'Incidencias', data['incidents'])
            # Users sheet
            self._add_sheet(wb, 'Usuarios', data['users'])
            
            # Remove default sheet if empty
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output
        else:
            raise ValueError(f"Unsupported format: {format}")

    def _add_sheet(self, wb, name: str, data: List[Dict]):
        """Add a sheet to workbook"""
        ws = wb.create_sheet(title=name)
        if not data:
            ws.append(["No data"])
            return
        
        headers = list(data[0].keys())
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row_idx, item in enumerate(data, 2):
            for col_idx, key in enumerate(headers, 1):
                value = item.get(key, '')
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=value)
