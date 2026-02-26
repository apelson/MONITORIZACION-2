"""
WatchTower by Siempria API - Main Server
Refactored version with modular routing
"""
from fastapi import FastAPI, APIRouter, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from typing import Optional
import asyncio
import os
import io
import ssl
import base64
import urllib.request
import uuid

# PDF/Excel exports
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm, inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image

# Local imports
from config import (
    devices_collection, organizations_collection, groups_collection,
    device_types_collection, public_dashboards_collection, logger
)
from services.auth_service import get_current_user, require_role, get_password_hash
from services.device_service import check_single_device, check_all_devices
from models import PublicDashboardConfig

# Import routers
from routes.auth import router as auth_router
from routes.users import router as users_router
from routes.organizations import router as organizations_router
from routes.devices import router as devices_router
from routes.settings import router as settings_router
from routes.statistics import router as statistics_router
from routes.upload import router as upload_router
from routes.backup import router as backup_router
from routes.logs import router as logs_router
from routes.reports import router as reports_router
from routes.incidents import router as incidents_router
from routes.security import router as security_router
from routes.infrastructure import router as infrastructure_router
from routes.device_images import router as device_images_router
from routes.cra_events import router as cra_events_router
from routes.camera_stream import router as camera_stream_router
from routes.roles import router as roles_router
from routes.superadmin_integrated import router as superadmin_router

# Multi-tenant routers (SaaS)
from routes.tenant_auth import router as tenant_auth_router
from routes.tenant_devices import router as tenant_devices_router
from routes.superadmin import router as saas_superadmin_router
from routes.billing import router as billing_router

# Super Admin for main platform multi-tenancy
from routes.superadmin_tenants import router as superadmin_tenants_router
from routes.jira import router as jira_router

# WebSocket router
from routes.websocket import router as websocket_router

# AI and SLA routes
from routes.ai_analysis import router as ai_router
from routes.sla_reports import router as sla_reports_router
from routes.download_files import router as download_router
from routes.dahua import router as dahua_router
from routes.vpn import router as vpn_router
from routes.system_stats import router as system_stats_router
from services.websocket_service import websocket_manager
from services.device_service import set_websocket_manager
from services.dahua_service import dahua_service

# ============ SCHEDULER FOR DAILY REPORTS ============
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from apscheduler.triggers.interval import IntervalTrigger

scheduler = AsyncIOScheduler()

async def scheduled_daily_report():
    """Send daily report at scheduled time"""
    from services.report_service import send_daily_report, get_report_settings
    try:
        settings = await get_report_settings()
        if settings.get("daily_report_enabled"):
            recipients = settings.get("daily_report_recipients", [])
            if recipients:
                result = await send_daily_report(recipients=recipients, days=1)
                logger.info(f"Scheduled daily report: {result}")
    except Exception as e:
        logger.error(f"Error in scheduled daily report: {e}")

# ============ APP LIFECYCLE ============

async def init_default_data():
    """Initialize default data if empty"""
    # Default device types
    if await device_types_collection.count_documents({}) == 0:
        default_types = [
            {"id": "type-camera", "name": "Cámara", "icon": "camera", "color": "#3b82f6", "is_default": True},
            {"id": "type-nas", "name": "NAS", "icon": "database", "color": "#8b5cf6", "is_default": True},
            {"id": "type-switch", "name": "Switch", "icon": "network", "color": "#22c55e", "is_default": True},
            {"id": "type-router", "name": "Router", "icon": "router", "color": "#f59e0b", "is_default": True},
            {"id": "type-server", "name": "Servidor", "icon": "server", "color": "#ef4444", "is_default": True},
            {"id": "type-other", "name": "Otro", "icon": "box", "color": "#6b7280", "is_default": True},
        ]
        await device_types_collection.insert_many(default_types)
        logger.info("Default device types created")
    
    # Default admin user
    from config import users_collection
    if await users_collection.count_documents({}) == 0:
        admin = {
            "id": str(uuid.uuid4()), "username": "admin", "email": "admin@siempria.com",
            "password_hash": get_password_hash("admin123"), "role": "admin",
            "full_name": "Administrador", "is_active": True, "group_ids": [],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await users_collection.insert_one(admin)
        
        # Create additional default users
        default_users = [
            {"username": "operador", "email": "operador@siempria.com", "password": "operador", "role": "operator", "full_name": "Operador"},
            {"username": "tecnico", "email": "tecnico@siempria.com", "password": "tecnico123", "role": "technician", "full_name": "Técnico"},
        ]
        for u in default_users:
            user = {
                "id": str(uuid.uuid4()), "username": u["username"], "email": u["email"],
                "password_hash": get_password_hash(u["password"]), "role": u["role"],
                "full_name": u["full_name"], "is_active": True, "group_ids": [],
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await users_collection.insert_one(user)
        logger.info("Default users created")

async def periodic_device_check():
    """Background task for periodic device checks"""
    while True:
        try:
            await check_all_devices()
        except Exception as e:
            logger.error(f"Error in periodic check: {e}")
        await asyncio.sleep(120)  # Check every 2 minutes

async def periodic_dahua_check():
    """Scheduled task for periodic Dahua DVR/NVR checks - runs every 5 minutes"""
    logger.info("=" * 50)
    logger.info("[DAHUA SCHEDULER] Starting Dahua devices check")
    try:
        results = await dahua_service.check_all_devices()
        
        online_count = sum(1 for r in results if r.get("online"))
        offline_count = sum(1 for r in results if not r.get("online"))
        
        logger.info(f"[DAHUA SCHEDULER] Completed: {len(results)} devices checked")
        logger.info(f"[DAHUA SCHEDULER] Status: Online={online_count}, Offline={offline_count}")
        
        # Broadcast status update via WebSocket
        if websocket_manager:
            await websocket_manager.broadcast({
                "type": "dahua_status_update",
                "data": {
                    "total": len(results),
                    "online": online_count,
                    "offline": offline_count,
                    "timestamp": datetime.now(timezone.utc).isoformat()
                }
            })
    except Exception as e:
        logger.error(f"[DAHUA SCHEDULER] Error checking Dahua devices: {e}")
    logger.info("=" * 50)

async def periodic_vpn_check():
    """Scheduled task for periodic VPN checks - runs every 5 minutes
    Uses existing devices collection with VPN type"""
    logger.info("[VPN SCHEDULER] Starting VPN devices check")
    try:
        from config import devices_collection, device_types_collection
        
        # Get VPN device type ID
        vpn_type = await device_types_collection.find_one(
            {"$or": [
                {"name": {"$regex": "vpn", "$options": "i"}},
                {"name": {"$regex": "tunnel", "$options": "i"}}
            ]},
            {"_id": 0}
        )
        
        if not vpn_type:
            # Fallback: search by name containing VPN
            devices = await devices_collection.find(
                {"name": {"$regex": "vpn", "$options": "i"}},
                {"_id": 0}
            ).to_list(length=100)
        else:
            devices = await devices_collection.find(
                {"device_type_id": vpn_type.get("id")},
                {"_id": 0}
            ).to_list(length=100)
        
        online_count = sum(1 for d in devices if d.get("status") == "online")
        offline_count = len(devices) - online_count
        
        logger.info(f"[VPN SCHEDULER] Found {len(devices)} VPN devices, Online={online_count}, Offline={offline_count}")
        
        # Broadcast VPN status update via WebSocket
        if websocket_manager:
            await websocket_manager.broadcast({
                "type": "vpn_status_update",
                "data": {
                    "total": len(devices),
                    "online": online_count,
                    "offline": offline_count,
                    "timestamp": datetime.now(timezone.utc).isoformat()
                }
            })
    except Exception as e:
        logger.error(f"[VPN SCHEDULER] Error checking VPN devices: {e}")

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Create database indexes for performance
    from config import create_indexes
    await create_indexes()
    
    # Initialize WebSocket manager for real-time alerts
    set_websocket_manager(websocket_manager)
    logger.info("WebSocket manager initialized for real-time alerts")
    
    await init_default_data()
    asyncio.create_task(periodic_device_check())
    
    # Start scheduler for daily reports
    # Default: 8:00 AM every day (will be overridden by settings)
    scheduler.add_job(
        scheduled_daily_report,
        CronTrigger(hour=8, minute=0),
        id="daily_report",
        replace_existing=True
    )
    
    # Add Dahua devices check every 5 minutes
    scheduler.add_job(
        periodic_dahua_check,
        IntervalTrigger(minutes=5),
        id="dahua_check",
        replace_existing=True
    )
    logger.info("Dahua scheduler started - checking every 5 minutes")
    
    # Add VPN devices check every 5 minutes
    scheduler.add_job(
        periodic_vpn_check,
        IntervalTrigger(minutes=5),
        id="vpn_check",
        replace_existing=True
    )
    logger.info("VPN scheduler started - checking every 5 minutes")
    
    scheduler.start()
    logger.info("Scheduler started for daily reports")
    
    logger.info("WatchTower by Siempria API started - OPTIMIZED with WebSockets")
    yield
    scheduler.shutdown()
    logger.info("WatchTower by Siempria API stopped")

# ============ APP SETUP ============

app = FastAPI(title="WatchTower by Siempria API", version="3.0", lifespan=lifespan)
api_router = APIRouter(prefix="/api")

# Include all routers - Original (single tenant)
api_router.include_router(auth_router)
api_router.include_router(users_router)
api_router.include_router(organizations_router)
api_router.include_router(devices_router)
api_router.include_router(settings_router)
api_router.include_router(statistics_router)
api_router.include_router(upload_router)
api_router.include_router(backup_router)
api_router.include_router(logs_router)
api_router.include_router(reports_router)
api_router.include_router(incidents_router)
api_router.include_router(security_router)
api_router.include_router(infrastructure_router)
api_router.include_router(device_images_router)
api_router.include_router(cra_events_router)
api_router.include_router(camera_stream_router)
api_router.include_router(roles_router)
api_router.include_router(superadmin_router)

# Include multi-tenant routers (SaaS)
api_router.include_router(tenant_auth_router)
api_router.include_router(tenant_devices_router)
api_router.include_router(saas_superadmin_router)
api_router.include_router(billing_router)

# Include Super Admin for main platform multi-tenancy
api_router.include_router(superadmin_tenants_router)

# Include JIRA integration routes
api_router.include_router(jira_router)

# Include WebSocket router
api_router.include_router(websocket_router)

# Include AI and SLA routes
api_router.include_router(ai_router)
api_router.include_router(sla_reports_router)

# Include Dahua P2P routes
api_router.include_router(dahua_router)

# Include VPN monitoring routes
api_router.include_router(vpn_router)

# Include System Stats routes
api_router.include_router(system_stats_router)

# Include download router for production updates
api_router.include_router(download_router)

# ============ DOWNLOAD ENDPOINTS FOR PRODUCTION UPDATE ============

@api_router.get("/dl/noc")
async def dl_noc():
    """Download corrected NOCDashboard.jsx"""
    file_path = "/app/backend/static_files/NOCDashboard_corrected.jsx"
    if os.path.exists(file_path):
        return FileResponse(file_path, filename="NOCDashboard.jsx", media_type='text/plain; charset=utf-8')
    raise HTTPException(status_code=404, detail="File not found")

@api_router.get("/dl/vpn")
async def dl_vpn():
    """Download vpn.py"""
    file_path = "/app/backend/routes/vpn.py"
    if os.path.exists(file_path):
        return FileResponse(file_path, filename="vpn.py", media_type='text/plain; charset=utf-8')
    raise HTTPException(status_code=404, detail="File not found")

@api_router.get("/dl/system-stats")
async def dl_system_stats():
    """Download system_stats.py"""
    file_path = "/app/backend/routes/system_stats.py"
    if os.path.exists(file_path):
        return FileResponse(file_path, filename="system_stats.py", media_type='text/plain; charset=utf-8')
    raise HTTPException(status_code=404, detail="File not found")

@api_router.get("/dl/mini-ecg")
async def dl_mini_ecg():
    """Download MiniECG.jsx"""
    file_path = "/app/frontend/src/components/common/MiniECG.jsx"
    if os.path.exists(file_path):
        return FileResponse(file_path, filename="MiniECG.jsx", media_type='text/plain; charset=utf-8')
    raise HTTPException(status_code=404, detail="File not found")

@api_router.get("/dl/system-monitor")
async def dl_system_monitor():
    """Download SystemResourceMonitor.jsx"""
    file_path = "/app/frontend/src/components/common/SystemResourceMonitor.jsx"
    if os.path.exists(file_path):
        return FileResponse(file_path, filename="SystemResourceMonitor.jsx", media_type='text/plain; charset=utf-8')
    raise HTTPException(status_code=404, detail="File not found")

@api_router.get("/dl/vpn-widget")
async def dl_vpn_widget():
    """Download VPNWidget.jsx"""
    file_path = "/app/frontend/src/components/noc/widgets/VPNWidget.jsx"
    if os.path.exists(file_path):
        return FileResponse(file_path, filename="VPNWidget.jsx", media_type='text/plain; charset=utf-8')
    raise HTTPException(status_code=404, detail="File not found")

@api_router.get("/dl/login")
async def dl_login():
    """Download LoginPage.jsx with spectacular logo"""
    file_path = "/app/frontend/src/components/auth/LoginPage.jsx"
    if os.path.exists(file_path):
        return FileResponse(file_path, filename="LoginPage.jsx", media_type='text/plain; charset=utf-8')
    raise HTTPException(status_code=404, detail="File not found")

@api_router.get("/dl/devices-py")
async def dl_devices_py():
    """Download devices.py with critical device fix"""
    file_path = "/app/backend/routes/devices.py"
    if os.path.exists(file_path):
        return FileResponse(file_path, filename="devices.py", media_type='text/plain; charset=utf-8')
    raise HTTPException(status_code=404, detail="File not found")

# ============ ROOT & IMAGE PROXY ============

@api_router.get("/")
async def root():
    return {"message": "WatchTower by Siempria API v3.1 (with WebSockets)"}

@api_router.get("/download-build")
async def download_build():
    """Download the frontend build tar.gz for production deployment"""
    import os
    build_path = os.path.join(os.path.dirname(__file__), "frontend_build_clean.tar.gz")
    if not os.path.exists(build_path):
        raise HTTPException(status_code=404, detail="Build file not found")
    
    def iterfile():
        with open(build_path, "rb") as f:
            yield from f
    
    return StreamingResponse(
        iterfile(),
        media_type="application/gzip",
        headers={"Content-Disposition": "attachment; filename=frontend_build.tar.gz"}
    )

@api_router.get("/download-file")
async def download_file(path: str):
    """Download a file for production deployment"""
    import os
    allowed_files = {
        "camera_stream.py": "/app/backend/routes/camera_stream.py",
        "devices.py": "/app/backend/routes/devices.py", 
        "cra_events.py": "/app/backend/routes/cra_events.py",
        "roles.py": "/app/backend/routes/roles.py",
        "superadmin_integrated.py": "/app/backend/routes/superadmin_integrated.py",
        "device_service.py": "/app/backend/services/device_service.py",
        "config.py": "/app/backend/config.py",
        "server.py": "/app/backend/server.py",
        "CRADashboard.jsx": "/app/frontend/src/components/panels/CRADashboard.jsx",
        "LiveViewer.jsx": "/app/frontend/src/components/panels/LiveViewer.jsx",
        "AlertsPanel.jsx": "/app/frontend/src/components/panels/AlertsPanel.jsx",
        "StatisticsPanel.jsx": "/app/frontend/src/components/panels/StatisticsPanel.jsx",
        "IncidentsPanel.jsx": "/app/frontend/src/components/panels/IncidentsPanel.jsx",
        "InfrastructurePanel.jsx": "/app/frontend/src/components/panels/InfrastructurePanel.jsx",
        "ServerCard.jsx": "/app/frontend/src/components/devices/ServerCard.jsx",
        "RolesManager.jsx": "/app/frontend/src/components/settings/RolesManager.jsx",
        "SuperAdminTab.jsx": "/app/frontend/src/components/settings/SuperAdminTab.jsx",
        "LiveViewerFloatingButton.jsx": "/app/frontend/src/components/common/LiveViewerFloatingButton.jsx",
        "SectionLoader.jsx": "/app/frontend/src/components/common/SectionLoader.jsx",
        "App.js": "/app/frontend/src/App.js",
        "NOCDashboard.jsx": "/app/frontend/src/components/panels/NOCDashboard.jsx",
        "NOCFloatingButton.jsx": "/app/frontend/src/components/common/NOCFloatingButton.jsx",
        "alerts_index.js": "/app/frontend/src/components/alerts/index.js",
        "DeviceStatusGrid.jsx": "/app/frontend/src/components/alerts/DeviceStatusGrid.jsx",
        "DeviceHistoryModal.jsx": "/app/frontend/src/components/alerts/DeviceHistoryModal.jsx",
        "AlertBell.jsx": "/app/frontend/src/components/alerts/AlertBell.jsx",
        "translation_es.json": "/app/frontend/src/locales/es/translation.json",
        "translation_de.json": "/app/frontend/src/locales/de/translation.json",
        "favicon.ico": "/app/frontend/public/favicon.ico",
        "logo192.png": "/app/frontend/public/logo192.png",
        "logo512.png": "/app/frontend/public/logo512.png",
        "favicon-16x16.png": "/app/frontend/public/icons/favicon-16x16.png",
        "favicon-32x32.png": "/app/frontend/public/icons/favicon-32x32.png",
        "icon-72x72.png": "/app/frontend/public/icons/icon-72x72.png",
        "icon-96x96.png": "/app/frontend/public/icons/icon-96x96.png",
        "icon-128x128.png": "/app/frontend/public/icons/icon-128x128.png",
        "icon-144x144.png": "/app/frontend/public/icons/icon-144x144.png",
        "icon-152x152.png": "/app/frontend/public/icons/icon-152x152.png",
        "icon-192x192.png": "/app/frontend/public/icons/icon-192x192.png",
        "icon-384x384.png": "/app/frontend/public/icons/icon-384x384.png",
        "icon-512x512.png": "/app/frontend/public/icons/icon-512x512.png",
        # New NOC components
        "ReportSettings.jsx": "/app/frontend/src/components/settings/ReportSettings.jsx",
        "NOCStatsBar.jsx": "/app/frontend/src/components/noc/NOCStatsBar.jsx",
        "NOCCRASection.jsx": "/app/frontend/src/components/noc/NOCCRASection.jsx",
        "NOCHistorySection.jsx": "/app/frontend/src/components/noc/NOCHistorySection.jsx",
        "NOCAlertsSection.jsx": "/app/frontend/src/components/noc/NOCAlertsSection.jsx",
        "NOCMapSection.jsx": "/app/frontend/src/components/noc/NOCMapSection.jsx",
        "NOCOrganizationsSection.jsx": "/app/frontend/src/components/noc/NOCOrganizationsSection.jsx",
        "noc_index.js": "/app/frontend/src/components/noc/index.js",
        # AI and SLA components
        "AIInsightsPanel.jsx": "/app/frontend/src/components/settings/AIInsightsPanel.jsx",
        "SLAReportsPanel.jsx": "/app/frontend/src/components/settings/SLAReportsPanel.jsx",
        # Common components
        "SystemECG.jsx": "/app/frontend/src/components/common/SystemECG.jsx",
        "PWAInstallPrompt.jsx": "/app/frontend/src/components/common/PWAInstallPrompt.jsx",
        # Dashboard widgets
        "DashboardWidgets.jsx": "/app/frontend/src/components/dashboard/DashboardWidgets.jsx",
        # NOC Widgets
        "widgets_index.js": "/app/frontend/src/components/noc/widgets/index.js",
        "StatsWidget.jsx": "/app/frontend/src/components/noc/widgets/StatsWidget.jsx",
        "UptimeWidget.jsx": "/app/frontend/src/components/noc/widgets/UptimeWidget.jsx",
        "SystemMonitorWidget.jsx": "/app/frontend/src/components/noc/widgets/SystemMonitorWidget.jsx",
        "CRAWidget.jsx": "/app/frontend/src/components/noc/widgets/CRAWidget.jsx",
        "OrganizationsWidget.jsx": "/app/frontend/src/components/noc/widgets/OrganizationsWidget.jsx",
        "OfflineWidget.jsx": "/app/frontend/src/components/noc/widgets/OfflineWidget.jsx",
        "HistoryWidget.jsx": "/app/frontend/src/components/noc/widgets/HistoryWidget.jsx",
        "AlertsWidget.jsx": "/app/frontend/src/components/noc/widgets/AlertsWidget.jsx",
        # NOC Refactored components
        "NOCDashboardRefactored.jsx": "/app/frontend/src/components/panels/NOCDashboardRefactored.jsx",
        "DraggableGrid.jsx": "/app/frontend/src/components/noc/DraggableGrid.jsx",
        "DashboardFilters.jsx": "/app/frontend/src/components/noc/DashboardFilters.jsx",
        "NOCHeader.jsx": "/app/frontend/src/components/noc/NOCHeader.jsx",
        # Backend routes
        "users.py": "/app/backend/routes/users.py",
        "auth.py": "/app/backend/routes/auth.py",
    }
    
    if path not in allowed_files:
        raise HTTPException(status_code=404, detail=f"File not found. Available: {list(allowed_files.keys())}")
    
    file_path = allowed_files[path]
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found on server")
    
    def iterfile():
        with open(file_path, "rb") as f:
            yield from f
    
    return StreamingResponse(
        iterfile(),
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename={path}"}
    )

@api_router.get("/image-proxy/{device_id}")
async def image_proxy(device_id: str, current_user: dict = Depends(get_current_user)):
    """Proxy to load device images with authentication"""
    device = await devices_collection.find_one({"id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Dispositivo no encontrado")
    
    protocol = device.get("camera_protocol", "http")
    ip = device.get("ip_address", "")
    port = device.get("port", 80)
    camera_user = device.get("camera_user", "")
    camera_password = device.get("camera_password", "")
    camera_path = device.get("camera_path", "")
    
    ssl_context = ssl.create_default_context()
    ssl_context.check_hostname = False
    ssl_context.verify_mode = ssl.CERT_NONE
    
    if not camera_path:
        image_url = device.get("image_url", "")
        if not image_url:
            raise HTTPException(status_code=404, detail="No hay imagen configurada")
        try:
            request = urllib.request.Request(image_url)
            request.add_header("User-Agent", "Mozilla/5.0 SiempriaMonitor/1.0")
            with urllib.request.urlopen(request, timeout=10, context=ssl_context) as response:
                image_data = response.read()
                content_type = response.headers.get('Content-Type', 'image/jpeg')
            return StreamingResponse(io.BytesIO(image_data), media_type=content_type, headers={"Cache-Control": "max-age=30"})
        except Exception as e:
            logger.error(f"Error fetching image for device {device_id}: {str(e)}")
            raise HTTPException(status_code=502, detail="No se pudo cargar la imagen")
    
    clean_url = f"{protocol}://{ip}:{port}{camera_path}"
    
    try:
        request = urllib.request.Request(clean_url)
        if camera_user and camera_password:
            credentials = base64.b64encode(f"{camera_user}:{camera_password}".encode()).decode()
            request.add_header("Authorization", f"Basic {credentials}")
        request.add_header("User-Agent", "Mozilla/5.0 SiempriaMonitor/1.0")
        
        with urllib.request.urlopen(request, timeout=10, context=ssl_context) as response:
            image_data = response.read()
            content_type = response.headers.get('Content-Type', 'image/jpeg')
        
        return StreamingResponse(io.BytesIO(image_data), media_type=content_type, headers={"Cache-Control": "max-age=30"})
    except urllib.error.HTTPError as e:
        raise HTTPException(status_code=502, detail=f"Error al cargar imagen: HTTP {e.code}")
    except urllib.error.URLError as e:
        raise HTTPException(status_code=502, detail="No se pudo conectar al dispositivo")
    except Exception as e:
        raise HTTPException(status_code=500, detail="Error al cargar imagen")

# ============ PUBLIC DASHBOARD ============

@api_router.get("/organizations/{org_id}/public-dashboard")
async def get_public_dashboard_config(org_id: str, current_user: dict = Depends(require_role(["admin"]))):
    config = await public_dashboards_collection.find_one({"organization_id": org_id}, {"_id": 0})
    return {"config": config or {"enabled": False, "password": None, "show_images": True, "show_details": False}}

@api_router.post("/organizations/{org_id}/public-dashboard")
async def save_public_dashboard_config(org_id: str, config: PublicDashboardConfig, current_user: dict = Depends(require_role(["admin"]))):
    existing = await public_dashboards_collection.find_one({"organization_id": org_id})
    public_token = existing.get("public_token") if existing else str(uuid.uuid4())
    
    update_data = {"organization_id": org_id, "public_token": public_token, **config.model_dump()}
    await public_dashboards_collection.update_one({"organization_id": org_id}, {"$set": update_data}, upsert=True)
    return {"message": "Configuración guardada", "public_url": f"/public/{public_token}"}

@api_router.get("/public/{token}")
async def get_public_dashboard(token: str, password: Optional[str] = None):
    """Get public dashboard data (no auth required)"""
    config = await public_dashboards_collection.find_one({"public_token": token}, {"_id": 0})
    if not config:
        raise HTTPException(status_code=404, detail="Dashboard no encontrado")
    if not config.get("enabled"):
        raise HTTPException(status_code=403, detail="Dashboard público deshabilitado")
    if config.get("password") and config.get("password") != password:
        raise HTTPException(status_code=401, detail="Contraseña incorrecta")
    
    org_id = config.get("organization_id")
    org = await organizations_collection.find_one({"id": org_id}, {"_id": 0})
    if not org:
        raise HTTPException(status_code=404, detail="Organización no encontrada")
    
    groups = await groups_collection.find({"organization_id": org_id}, {"_id": 0}).to_list(length=100)
    group_ids = [g["id"] for g in groups]
    devices = await devices_collection.find({"group_id": {"$in": group_ids}}, {"_id": 0}).to_list(length=500)
    
    show_details = config.get("show_details", False)
    show_images = config.get("show_images", True)
    
    filtered_devices = []
    for device in devices:
        d = {
            "id": device.get("id"), "name": device.get("name"), "status": device.get("status"),
            "last_check": device.get("last_check"), "device_type_id": device.get("device_type_id"),
            "group_id": device.get("group_id"), "brand": device.get("brand"),
            "model": device.get("model"), "location": device.get("location")
        }
        if show_details:
            d["ip_address"] = device.get("ip_address")
            d["port"] = device.get("port")
        if show_images and device.get("camera_path"):
            d["has_image"] = True
        filtered_devices.append(d)
    
    device_types = await device_types_collection.find({}, {"_id": 0}).to_list(length=50)
    total = len(filtered_devices)
    online = len([d for d in filtered_devices if d.get("status") == "online"])
    offline = len([d for d in filtered_devices if d.get("status") == "offline"])
    
    return {
        "organization": {"name": org.get("name"), "logo_url": org.get("logo_url")},
        "groups": groups, "devices": filtered_devices, "device_types": device_types,
        "stats": {"total": total, "online": online, "offline": offline, "uptime_percent": round((online / total * 100) if total > 0 else 0, 1)},
        "config": {"show_images": show_images, "show_details": show_details, "requires_password": bool(config.get("password"))}
    }

# ============ EXPORT ROUTES ============

@api_router.get("/export/excel")
async def export_excel(organization_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Export devices to Excel file"""
    devices = await devices_collection.find({}, {"_id": 0}).to_list(length=None)
    organizations = await organizations_collection.find({}, {"_id": 0}).to_list(length=None)
    groups = await groups_collection.find({}, {"_id": 0}).to_list(length=None)
    device_types = await device_types_collection.find({}, {"_id": 0}).to_list(length=None)
    
    if organization_id:
        org_group_ids = [g["id"] for g in groups if g.get("organization_id") == organization_id]
        devices = [d for d in devices if d.get("group_id") in org_group_ids]
    
    org_dict = {o["id"]: o for o in organizations}
    group_dict = {g["id"]: g for g in groups}
    type_dict = {t["id"]: t for t in device_types}
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Dispositivos"
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers = ["Nombre", "IP:Puerto", "Estado", "Tipo", "Organización", "Grupo", "Marca", "Modelo", "Ubicación"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_num, device in enumerate(devices, 2):
        group = group_dict.get(device.get("group_id"))
        org = org_dict.get(group.get("organization_id")) if group else None
        device_type = type_dict.get(device.get("device_type_id"))
        
        row_data = [
            device.get("name", ""),
            f"{device.get('ip_address', '')}:{device.get('port', '')}",
            "Online" if device.get("status") == "online" else "Offline" if device.get("status") == "offline" else "?",
            device_type.get("name", "") if device_type else "",
            org.get("name", "") if org else "",
            group.get("name", "") if group else "",
            device.get("brand", ""),
            device.get("model", ""),
            device.get("location", "")
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            if col == 3:
                if value == "Online":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif value == "Offline":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    column_widths = [25, 20, 12, 15, 20, 20, 15, 20, 25]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws.freeze_panes = "A2"
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"dispositivos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           headers={"Content-Disposition": f"attachment; filename={filename}"})

# ============ MOBOTIX INFO ENDPOINT ============

@api_router.get("/devices/{device_id}/mobotix-info")
async def get_mobotix_info(device_id: str, current_user: dict = Depends(get_current_user)):
    """Get Mobotix camera information using HTTP API - Optimized with parallel requests"""
    import re
    
    device = await devices_collection.find_one({"id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Dispositivo no encontrado")
    
    protocol = device.get("camera_protocol", "http")
    ip = device.get("ip_address", "")
    port = device.get("port", 80)
    camera_user = device.get("camera_user", "")
    camera_password = device.get("camera_password", "")
    
    if not ip:
        raise HTTPException(status_code=400, detail="IP del dispositivo no configurada")
    
    base_url = f"{protocol}://{ip}:{port}"
    info = {
        "device_id": device_id, 
        "device_name": device.get("name", ""),
        "ip_address": f"{ip}:{port}", 
        "protocol": protocol,
        "system": {},
        "networking": {},
        "storage": {},
        "sensors": {},
        "image": {},
        "recording": {},
        "raw_html": None,
        "errors": []
    }
    
    def make_request(url_path: str, timeout: int = 5):
        try:
            full_url = f"{base_url}{url_path}"
            import ssl
            ctx = ssl.create_default_context()
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE
            
            request = urllib.request.Request(full_url)
            if camera_user and camera_password:
                credentials = base64.b64encode(f"{camera_user}:{camera_password}".encode()).decode()
                request.add_header("Authorization", f"Basic {credentials}")
            request.add_header("User-Agent", "SiempriaMonitor/1.0")
            with urllib.request.urlopen(request, timeout=timeout, context=ctx) as response:
                return response.read().decode('utf-8', errors='ignore')
        except urllib.error.HTTPError as e:
            return f"HTTP_ERROR:{e.code}"
        except urllib.error.URLError as e:
            return f"URL_ERROR:{str(e.reason)}"
        except Exception as e:
            return f"ERROR:{str(e)}"
    
    def parse_camerainfo_html(html: str) -> dict:
        """Parse the /control/camerainfo HTML response to extract structured data"""
        import re
        import html as html_module
        
        def decode_html(text: str) -> str:
            """Decode HTML entities like &deg; to proper characters"""
            if text:
                return html_module.unescape(text)
            return text
        
        data = {
            "system": {},
            "networking": {},
            "storage": {},
            "sensors": {},
            "image": {},
            "recording": {},
            "firmware_version": None
        }
        
        # Extract firmware from JavaScript variable
        firmware_match = re.search(r'filesystem__version="([^"]+)"', html)
        if firmware_match:
            data["firmware_version"] = firmware_match.group(1)
        
        # Parse table rows - pattern: <td>Label</td>\n<td colspan=2>Value</td>
        # System section
        patterns = {
            "model": r'<td>Model</td>\s*<td[^>]*>([^<]+)</td>',
            "serial_number": r'<td>Serial Number</td>\s*<td[^>]*>([^<]+)</td>',
            "hardware": r'<td>Hardware</td>\s*<td[^>]*>([^<]+)</td>',
            "image_sensor": r'<td>Image Sensor</td>\s*<td[^>]*>([^<]+)</td>',
            "software": r'<td>Software</td>\s*<td[^>]*>([^<]+)</td>',
            "uptime": r'<td>Current Uptime</td>\s*<td[^>]*>([^<]+)</td>',
            "date_time": r'<td>Date and Time</td>\s*<td[^>]*>([^<\n]+)',
        }
        for key, pattern in patterns.items():
            match = re.search(pattern, html, re.IGNORECASE)
            if match:
                data["system"][key] = decode_html(match.group(1).strip())
        
        # Networking section
        net_patterns = {
            "camera_name": r'<td>Camera Name</td>\s*<td[^>]*>([^<]+)</td>',
            "ip_address": r'<td>IP Address</td>\s*<td[^>]*>([^<]+)</td>',
            "network_mask": r'<td>Network Mask</td>\s*<td[^>]*>([^<]+)</td>',
            "link_speed": r'<td>Link Speed and Duplex</td>\s*<td[^>]*>([^<]+)</td>',
        }
        for key, pattern in net_patterns.items():
            match = re.search(pattern, html, re.IGNORECASE)
            if match:
                data["networking"][key] = decode_html(match.group(1).strip())
        
        # Storage section
        storage_patterns = {
            "type": r'<tbody ID=[\'"]fileserver[\'"]>.*?<td>Type</td>\s*<td[^>]*>([^<]+)</td>',
            "flash_wear": r'<td>Flash Wear</td>\s*<td>([^<]+)</td>',
            "current_usage": r'<td>Current Usage</td>\s*<td[^>]*>([^<]+)</td>',
            "maximum_size": r'<td>Maximum Size</td>\s*<td[^>]*>([^<]+)</td>',
            "sequences": r'<td>Sequences </td>\s*<td[^>]*>([^<]+)</td>',
        }
        for key, pattern in storage_patterns.items():
            match = re.search(pattern, html, re.IGNORECASE | re.DOTALL)
            if match:
                data["storage"][key] = decode_html(match.group(1).strip())
        
        # Sensors section
        sensor_patterns = {
            "illumination": r'<td>Illumination</td>\s*<td[^>]*>([^<]+)</td>',
            "temperature": r'<td>Camera Temperature</td>\s*<td[^>]*>([^<]+)</td>',
        }
        for key, pattern in sensor_patterns.items():
            match = re.search(pattern, html, re.IGNORECASE)
            if match:
                data["sensors"][key] = decode_html(match.group(1).strip())
        
        # Image section
        image_patterns = {
            "video_codec": r'<td>Video Codec</td>\s*<td[^>]*>([^<]+)</td>',
            "image_quality": r'<td>Image Quality</td>\s*<td[^>]*>([^<]+)</td>',
            "image_properties": r'<td>Image Properties</td>\s*<td[^>]*>([^<]+)',
            "frame_rate": r'<td>Current Frame Rate</td>\s*<td[^>]*>([^<]+)</td>',
        }
        for key, pattern in image_patterns.items():
            match = re.search(pattern, html, re.IGNORECASE)
            if match:
                data["image"][key] = decode_html(match.group(1).strip())
        
        # Recording section
        rec_patterns = {
            "recording_mode": r'<td>Recording Mode</td>\s*<td[^>]*>([^<]+)</td>',
            "event_frame_rate": r'<td>Event Frame Rate</td>\s*<td[^>]*>([^<]+)</td>',
            "recording_status": r'<td>Recording</td>\s*<td[^>]*>([^<]+)</td>',
            "storage_state": r'<td>Storage State</td>\s*<td[^>]*>([^<]+)</td>',
        }
        for key, pattern in rec_patterns.items():
            match = re.search(pattern, html, re.IGNORECASE)
            if match:
                data["recording"][key] = decode_html(match.group(1).strip())
        
        # Time/NTP section
        data["time"] = {}
        time_patterns = {
            "time_server": r'<td>Time Server</td>\s*<td[^>]*>([^<]+)</td>',
            "ntp_server": r'<td>NTP Server</td>\s*<td[^>]*>([^<]+)</td>',
            "time_zone": r'<td>Time Zone</td>\s*<td[^>]*>([^<]+)</td>',
            "time_source": r'<td>Time Source</td>\s*<td[^>]*>([^<]+)</td>',
            "time_servers_protocol": r'<td>Time Servers</td>\s*<td[^>]*>([^<]+)</td>',
        }
        for key, pattern in time_patterns.items():
            match = re.search(pattern, html, re.IGNORECASE)
            if match:
                data["time"][key] = decode_html(match.group(1).strip())
        
        # Look for NTP pool servers (like 0.pool.ntp.org, 1.pool.ntp.org)
        ntp_pool_matches = re.findall(r'(\d+\.pool\.ntp\.org|ntp\d*\.[a-zA-Z0-9.-]+|time\.[a-zA-Z0-9.-]+)', html, re.IGNORECASE)
        if ntp_pool_matches:
            data["time"]["ntp_servers_found"] = list(set(ntp_pool_matches))
        
        # Also look for any IP addresses that might be NTP servers in time-related context
        # Search for time server IPs in input fields
        ntp_input_matches = re.findall(r'name=["\']?(?:ntp|time)[^"\']*["\']?\s+value=["\']?([^"\'>\s]+)', html, re.IGNORECASE)
        if ntp_input_matches:
            data["time"]["ntp_configured_servers"] = [s for s in ntp_input_matches if s and s.strip()]
        
        # Error/Alarm section
        data["alarms"] = {}
        alarm_patterns = {
            "active_alarms": r'<td>Active Alarms</td>\s*<td[^>]*>([^<]+)</td>',
            "error_messages": r'<td>Error Messages</td>\s*<td[^>]*>([^<]+)</td>',
            "system_messages": r'<td>System Messages</td>\s*<td[^>]*>([^<]+)</td>',
        }
        for key, pattern in alarm_patterns.items():
            match = re.search(pattern, html, re.IGNORECASE)
            if match:
                data["alarms"][key] = decode_html(match.group(1).strip())
        
        return data
    
    loop = asyncio.get_event_loop()
    
    # Run both requests in PARALLEL for faster response
    import concurrent.futures
    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        future_camerainfo = executor.submit(make_request, "/control/camerainfo")
        future_time = executor.submit(make_request, "/admin/time")
        
        result = future_camerainfo.result()
        time_config_result = future_time.result()
    
    # Process NTP from admin page
    ntp_servers_from_admin = []
    if time_config_result and not time_config_result.startswith(("HTTP_ERROR", "URL_ERROR", "ERROR")):
        ntp_matches = re.findall(r'value=["\']?(\d+\.pool\.ntp\.org|[a-zA-Z0-9.-]+\.ntp\.[a-zA-Z]+|ntp\.[a-zA-Z0-9.-]+)', time_config_result, re.IGNORECASE)
        ntp_servers_from_admin = list(set(ntp_matches))
        if re.search(r'selected[^>]*>NTP<|value=["\']NTP["\'][^>]*selected|checked[^>]*NTP', time_config_result, re.IGNORECASE):
            info["ntp_protocol_enabled"] = True
    
    if not result.startswith(("HTTP_ERROR", "URL_ERROR", "ERROR")):
        parsed = parse_camerainfo_html(result)
        info["system"] = parsed["system"]
        info["networking"] = parsed["networking"]
        info["storage"] = parsed["storage"]
        info["sensors"] = parsed["sensors"]
        info["image"] = parsed["image"]
        info["recording"] = parsed["recording"]
        info["firmware_version"] = parsed["firmware_version"]
        info["time"] = parsed.get("time", {})
        info["alarms"] = parsed.get("alarms", {})
        
        # Add NTP servers found from admin page
        if ntp_servers_from_admin:
            if "ntp_servers_found" not in info["time"]:
                info["time"]["ntp_servers_found"] = []
            info["time"]["ntp_servers_found"].extend(ntp_servers_from_admin)
            info["time"]["ntp_servers_found"] = list(set(info["time"]["ntp_servers_found"]))
        
        # Determine recording status
        rec_mode = parsed["recording"].get("recording_mode", "").lower()
        rec_status = parsed["recording"].get("recording_status", "").lower()
        if "continuous" in rec_mode or "armed" in rec_mode or "recording" in rec_status or "active" in rec_status:
            info["is_recording"] = True
        elif "off" in rec_mode or "none" in rec_mode or "idle" in rec_status:
            info["is_recording"] = False
        else:
            info["is_recording"] = None  # Unknown
        
        # Determine NTP status - check multiple sources
        time_data = parsed.get("time", {})
        ntp = time_data.get("ntp_server") or time_data.get("time_server")
        ntp_servers_found = time_data.get("ntp_servers_found", [])
        ntp_configured_servers = time_data.get("ntp_configured_servers", [])
        time_protocol = time_data.get("time_servers_protocol", "")
        
        # Build list of all NTP servers found
        all_ntp_servers = []
        if ntp:
            all_ntp_servers.append(ntp)
        all_ntp_servers.extend(ntp_servers_found)
        all_ntp_servers.extend(ntp_configured_servers)
        all_ntp_servers = list(set([s for s in all_ntp_servers if s and s.strip() and s.lower() not in ["none", "-", ""]]))
        
        # Check if NTP is configured
        ntp_configured = bool(all_ntp_servers) or time_protocol.upper() == "NTP"
        
        info["ntp_server"] = ", ".join(all_ntp_servers) if all_ntp_servers else None
        info["ntp_configured"] = ntp_configured
        info["time"] = time_data  # Include full time data for frontend
        
        # Determine if there are errors
        alarms = parsed.get("alarms", {})
        active = alarms.get("active_alarms", "")
        errors = alarms.get("error_messages", "")
        info["has_errors"] = bool(
            (active and active.lower() not in ["none", "0", "-", ""]) or
            (errors and errors.lower() not in ["none", "0", "-", ""])
        )
        info["error_details"] = []
        if active and active.lower() not in ["none", "0", "-", ""]:
            info["error_details"].append(f"Alarmas: {active}")
        if errors and errors.lower() not in ["none", "0", "-", ""]:
            info["error_details"].append(f"Errores: {errors}")
        
        # Update device with firmware version if found
        if parsed["firmware_version"]:
            await devices_collection.update_one(
                {"id": device_id},
                {"$set": {
                    "firmware_version": parsed["firmware_version"],
                    "firmware_last_check": datetime.now(timezone.utc).isoformat(),
                    "camera_model": parsed["system"].get("model", ""),
                    "camera_sensor": parsed["system"].get("image_sensor", "")
                }}
            )
    else:
        info["errors"].append(f"camerainfo: {result}")
    
    return info

# ============ INCLUDE ROUTER & CORS ============

app.include_router(api_router)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"]
)

# Download route for App.js
from fastapi.responses import PlainTextResponse
import os

@app.get("/api/download/app-js")
async def download_app_js():
    file_path = "/app/frontend/src/App.js"
    if os.path.exists(file_path):
        with open(file_path, 'r') as f:
            content = f.read()
        return PlainTextResponse(content, media_type="text/plain")
    return PlainTextResponse("File not found", status_code=404)

@app.get("/api/download/server-py")
async def download_server_py():
    file_path = "/app/backend/server.py"
    if os.path.exists(file_path):
        with open(file_path, 'r') as f:
            content = f.read()
        return PlainTextResponse(content, media_type="text/plain")
    return PlainTextResponse("File not found", status_code=404)

@app.get("/api/download/file/{file_path:path}")
async def download_any_file(file_path: str):
    """Download any file from /app/"""
    # Try frontend first, then backend
    paths_to_try = [
        f"/app/frontend/src/{file_path}",
        f"/app/{file_path}"
    ]
    
    for full_path in paths_to_try:
        if os.path.exists(full_path):
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                return PlainTextResponse(content, media_type="text/plain; charset=utf-8")
            except:
                pass
    return PlainTextResponse("File not found", status_code=404)

@app.get("/api/download/saas-app")
async def download_saas_app():
    file_path = "/app/frontend/src/SaaSApp.jsx"
    if os.path.exists(file_path):
        with open(file_path, 'r') as f:
            content = f.read()
        return PlainTextResponse(content, media_type="text/plain")
    return PlainTextResponse("File not found", status_code=404)

@app.get("/api/download/index-js")
async def download_index_js():
    file_path = "/app/frontend/src/index.js"
    if os.path.exists(file_path):
        with open(file_path, 'r') as f:
            content = f.read()
        return PlainTextResponse(content, media_type="text/plain")
    return PlainTextResponse("File not found", status_code=404)

# Endpoint para descargar archivos de actualización
from fastapi.responses import FileResponse
import os

@app.get("/api/download/{filename}")
async def download_file(filename: str):
    file_path = f"/app/backend/static_files/{filename}"
    if os.path.exists(file_path):
        return FileResponse(file_path, filename=filename, media_type='application/octet-stream')
    raise HTTPException(status_code=404, detail=f"File not found: {filename}")

# Endpoint público para descargar actualización (sin auth)
@api_router.get("/download-update-package")
async def download_update_package_public():
    """Download siempria update package - no auth required"""
    file_path = "/app/backend/static_files/siempria_update.tar.gz"
    if os.path.exists(file_path):
        return FileResponse(file_path, filename="siempria_update.tar.gz", media_type='application/gzip')
    raise HTTPException(status_code=404, detail="Update package not found")

# Endpoint para descargar NOCDashboard corregido
@api_router.get("/download-noc")
async def download_noc_corrected():
    """Download corrected NOCDashboard.jsx - no auth required"""
    file_path = "/app/backend/static_files/NOCDashboard_corrected.jsx"
    if os.path.exists(file_path):
        return FileResponse(file_path, filename="NOCDashboard.jsx", media_type='text/plain')
    raise HTTPException(status_code=404, detail="File not found")

# Endpoint para descargar vpn.py
@api_router.get("/download-vpn")
async def download_vpn():
    """Download vpn.py - no auth required"""
    file_path = "/app/backend/routes/vpn.py"
    if os.path.exists(file_path):
        return FileResponse(file_path, filename="vpn.py", media_type='text/plain')
    raise HTTPException(status_code=404, detail="File not found")

# Endpoint para descargar system_stats.py
@api_router.get("/download-system-stats")
async def download_system_stats():
    """Download system_stats.py - no auth required"""
    file_path = "/app/backend/routes/system_stats.py"
    if os.path.exists(file_path):
        return FileResponse(file_path, filename="system_stats.py", media_type='text/plain')
    raise HTTPException(status_code=404, detail="File not found")

# Endpoint para descargar MiniECG.jsx
@api_router.get("/download-mini-ecg")
async def download_mini_ecg():
    """Download MiniECG.jsx - no auth required"""
    file_path = "/app/frontend/src/components/common/MiniECG.jsx"
    if os.path.exists(file_path):
        return FileResponse(file_path, filename="MiniECG.jsx", media_type='text/plain')
    raise HTTPException(status_code=404, detail="File not found")

# Endpoint para descargar SystemResourceMonitor.jsx
@api_router.get("/download-system-monitor")
async def download_system_monitor():
    """Download SystemResourceMonitor.jsx - no auth required"""
    file_path = "/app/frontend/src/components/common/SystemResourceMonitor.jsx"
    if os.path.exists(file_path):
        return FileResponse(file_path, filename="SystemResourceMonitor.jsx", media_type='text/plain')
    raise HTTPException(status_code=404, detail="File not found")

# Endpoint para descargar VPNWidget.jsx
@api_router.get("/download-vpn-widget")
async def download_vpn_widget():
    """Download VPNWidget.jsx - no auth required"""
    file_path = "/app/frontend/src/components/noc/widgets/VPNWidget.jsx"
    if os.path.exists(file_path):
        return FileResponse(file_path, filename="VPNWidget.jsx", media_type='text/plain')
    raise HTTPException(status_code=404, detail="File not found")

# Endpoint temporal para descargar App.js
@app.get("/api/download/appjs")
async def download_appjs():
    import os
    file_path = "/app/frontend/src/App.js"
    if os.path.exists(file_path):
        with open(file_path, 'r') as f:
            content = f.read()
        return {"content": content}
    return {"error": "File not found"}

# Endpoint para descargar archivos del frontend para producción
@app.get("/api/download-frontend/{filepath:path}")
async def download_frontend_file(filepath: str):
    """Download frontend files for production update"""
    import os
    
    # Map of allowed paths
    base_paths = {
        "components/noc": "/app/frontend/src/components/noc",
        "components/panels": "/app/frontend/src/components/panels", 
        "components/common": "/app/frontend/src/components/common",
        "routes": "/app/backend/routes",
        "services": "/app/backend/services",
    }
    
    # Find the file
    for prefix, base in base_paths.items():
        if filepath.startswith(prefix):
            full_path = f"/app/frontend/src/{filepath}"
            if os.path.exists(full_path):
                with open(full_path, 'r') as f:
                    return PlainTextResponse(f.read(), media_type="text/plain")
    
    # Try direct backend path
    backend_path = f"/app/backend/{filepath}"
    if os.path.exists(backend_path):
        with open(backend_path, 'r') as f:
            return PlainTextResponse(f.read(), media_type="text/plain")
            
    # Try direct frontend path
    frontend_path = f"/app/frontend/src/{filepath}"
    if os.path.exists(frontend_path):
        with open(frontend_path, 'r') as f:
            return PlainTextResponse(f.read(), media_type="text/plain")
    
    return PlainTextResponse(f"File not found: {filepath}", status_code=404)

# List all downloadable NOC files
@app.get("/api/list-noc-files")
async def list_noc_files():
    """List all NOC dashboard files available for download"""
    import os
    files = []
    
    # NOC components
    noc_path = "/app/frontend/src/components/noc"
    if os.path.exists(noc_path):
        for root, dirs, filenames in os.walk(noc_path):
            for f in filenames:
                if f.endswith('.jsx') or f.endswith('.js'):
                    rel_path = os.path.relpath(os.path.join(root, f), "/app/frontend/src")
                    files.append(rel_path)
    
    # NOC panels
    panels_path = "/app/frontend/src/components/panels"
    for f in ["NOCDashboard.jsx", "NOCDashboardRefactored.jsx"]:
        if os.path.exists(os.path.join(panels_path, f)):
            files.append(f"components/panels/{f}")
    
    # Backend websocket
    files.append("routes/websocket.py")
    
    return {"files": files}

# Temporary endpoint to download update files
@app.get("/download-update")
async def download_update():
    import os
    from fastapi.responses import FileResponse
    file_path = "/app/backend/static_files/siempria-update.tar.gz"
    if os.path.exists(file_path):
        return FileResponse(file_path, filename="siempria-update.tar.gz", media_type="application/gzip")
    return {"error": "File not found"}
